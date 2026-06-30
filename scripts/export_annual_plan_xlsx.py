import json
import math
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, InlineFont, TextBlock
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MONTHS = {
    1: "OCAK",
    2: "ŞUBAT",
    3: "MART",
    4: "NİSAN",
    5: "MAYIS",
    6: "HAZİRAN",
    7: "TEMMUZ",
    8: "AĞUSTOS",
    9: "EYLÜL",
    10: "EKİM",
    11: "KASIM",
    12: "ARALIK",
}


def tr_upper(value):
    text = str(value or "")
    replacements = {
        "i": "İ",
        "ı": "I",
        "ğ": "Ğ",
        "ü": "Ü",
        "ş": "Ş",
        "ö": "Ö",
        "ç": "Ç",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text.upper()


def parse_date(value):
    try:
        return datetime.strptime(value or "", "%Y-%m-%d")
    except ValueError:
        return None


def date_label(start, end):
    start_dt = parse_date(start)
    end_dt = parse_date(end)
    if not start_dt or not end_dt:
        return f"{start or ''} - {end or ''}".strip(" -")
    
    TurkishMonths = [
        "", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
        "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"
    ]
    
    start_day = f"{start_dt.day:02d}"
    end_day = f"{end_dt.day:02d}"
    
    if start_dt.year == end_dt.year:
        if start_dt.month == end_dt.month:
            return f"{start_day}-{end_day} {TurkishMonths[start_dt.month]} {start_dt.year}"
        else:
            return f"{start_day} {TurkishMonths[start_dt.month]} - {end_day} {TurkishMonths[end_dt.month]} {start_dt.year}"
    else:
        return f"{start_day} {TurkishMonths[start_dt.month]} {start_dt.year} - {end_day} {TurkishMonths[end_dt.month]} {end_dt.year}"


def week_date_label(week_no, start, end):
    week_text = f"{week_no}. Hafta" if week_no else "Hafta"
    date_text = date_label(start, end)
    return f"{week_text}\n{date_text}" if date_text else week_text


def month_label(start):
    start_dt = parse_date(start)
    return MONTHS[start_dt.month] if start_dt else ""


def clean(value):
    return re.sub(r"\s+\n", "\n", str(value or "").strip())


def rich_note_text(notes, exam_labels):
    labels = set(exam_labels or [])
    if not labels or not notes:
        return notes

    rich_text = CellRichText()
    bold_font = InlineFont(b=True)
    for index, line in enumerate(str(notes).splitlines()):
        text = line if index == 0 else f"\n{line}"
        if line in labels:
            rich_text.append(TextBlock(bold_font, text))
        else:
            rich_text.append(text)
    return rich_text


def topic_key(value):
    value = str(value or "").lower()
    replacements = {
        "ı": "i", "ş": "s", "ğ": "g", "ç": "c", "ö": "o", "ü": "u",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def clean_topic_value(value):
    value = clean(value)
    value = re.sub(r"\b(Sa)\s+(?=[yY\?])", r"\1", value)
    value = re.sub(r"(?<=[a-zçğıöşü])(?=[A-ZÇĞİÖŞÜ])", " ", value)
    value = re.sub(r"(?<=\?)(?=[A-ZÇĞİÖŞÜ])", " ", value)
    value = re.sub(r"\b([A-Za-zÇĞİÖŞÜçğıöşü])\s+([A-Za-zÇĞİÖŞÜçğıöşü]{2,})\b", r"\1\2", value)
    value = re.split(r"\b(?:Aşağıdaki|Asagidaki|Aşağıda|Asagida|A.{0,8}daki)\b", value, maxsplit=1, flags=re.I)[0]
    value = re.sub(r"(?<=[a-zçğıöşü])(?=[A-ZÇĞİÖŞÜ])", " ", value)
    value = re.sub(r"(?<=\?)(?=[A-ZÇĞİÖŞÜ])", " ", value)
    value = re.sub(r"\b(Sa)\s+(?=[yY\?])", r"\1", value)
    value = re.sub(r"\s+\d+\s*$", "", value)
    return clean(value)


def is_topic_excluded(value):
    key = topic_key(value)
    if not key:
        return True
    excluded = (
        "ders ici etkinlik", "ders ici", "etkinlik", "olcme ve degerlendirme",
        "olcme degerlendirme", "cevap anahtari", "alistirma", "uygulama",
        "sira sizde", "asagidaki", "asagida", "yapiniz", "yap n z", "cozunuz",
        "coz n z", "cevaplayiniz", "cevaplay n z", "inceleyiniz", "inceley n z",
        "ornek soru", "ornek uygulama",
    )
    if any(bit in key for bit in excluded):
        return True
    if re.match(r"^\d+\s+sa$", key):
        return True
    if re.match(r"^\d+\s+[a-z]{1,2}$", key):
        return True
    return False


def topic_text(item, show_title=True):
    pieces = []
    title = clean(item.get("title"))
    topics = item.get("topics") if isinstance(item.get("topics"), list) else []
    if title and show_title:
        pieces.append(tr_upper(f"Öğrenme Birimi: {title}"))
    if topics:
        for idx, topic in enumerate(topics):
            val = clean_topic_value(topic)
            if val:
                if is_topic_excluded(val):
                    continue
                if re.match(r"^((?:\d+(?:\.\d+)*|[A-Za-zÇĞİÖŞÜçğıöşü]|[IVXLCDMivxlcdm]+)\b\.?\s*|^[•\-*])", val):
                    pieces.append(val)
                else:
                    pieces.append(f"{idx + 1}. {val}")
    return "\n".join(pieces) if pieces else title


def outcome_text(items):
    outcomes = []
    for item in items:
        text = clean(item.get("outcomes"))
        if text and text not in outcomes:
            outcomes.append(text)
    return "\n".join(outcomes)


def build_rows(plan):
    rows = []
    last_unit_title = ""
    for week in plan.get("weeks", []):
        if week.get("skipped"):
            week_no = week.get("no", "")
            skip_note = clean(week.get("skipNote")) or "Ara / tatil"
            skip_range = date_label(week.get("start"), week.get("end"))
            skip_text = f"{skip_note} ({skip_range})" if skip_range else skip_note
            rows.append({
                "kind": "break",
                "text": skip_text,
                "weekDate": week_date_label(week_no, week.get("start"), week.get("end")),
            })
            continue
        items = week.get("items") or []
        
        formatted_topics = []
        for item in items:
            item_title = clean(item.get("title"))
            show_title = False
            if item_title != last_unit_title:
                show_title = True
                last_unit_title = item_title
            
            t_text = topic_text(item, show_title=show_title)
            if t_text:
                formatted_topics.append(t_text)
                
        rows.append({
            "kind": "lesson",
            "weekDate": week_date_label(week.get("no", ""), week.get("start"), week.get("end")),
            "hours": sum(int(item.get("hours") or 0) for item in items) or int(plan.get("weeklyHours") or 0),
            "outcomes": outcome_text(items),
            "topics": "\n\n".join(formatted_topics),
            "methods": clean(plan.get("methods")) or "Anlatım, soru-cevap, uygulamalı gösteri, araştırma, uygulama, bireysel öğrenme, tartışma",
            "materials": clean(plan.get("materials")) or "Akıllı tahta, ders kitabı, çalışma kağıtları, örnek uygulamalar",
            "notes": clean(week.get("note") or ""),
            "examLabels": week.get("examLabels") or [],
        })
    return rows


def display_length(value):
    text = str(value or "")
    # Turkish text is mostly single-width, but punctuation and numbers are
    # compact in Excel. This keeps the estimate practical without overpacking.
    return max(1, int(len(text) * 0.95))


def merged_cell_width(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            if cell.row != merged_range.min_row or cell.column != merged_range.min_col:
                return 0
            return sum(
                ws.column_dimensions[get_column_letter(col)].width or 10
                for col in range(merged_range.min_col, merged_range.max_col + 1)
            )
    return ws.column_dimensions[get_column_letter(cell.column)].width or 10


def wrapped_line_count(text, width):
    chars_per_line = max(8, int(width * 0.92))
    lines = 0
    for raw_line in str(text or "").splitlines() or [""]:
        line = raw_line.strip()
        if not line:
            lines += 1
            continue
        # Very long topic numbers/titles can wrap mid-token in Excel.
        lines += max(1, math.ceil(display_length(line) / chars_per_line))
    return lines


def autosize_rows(ws):
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                width = merged_cell_width(ws, cell)
                if width <= 0:
                    continue
                max_lines = max(max_lines, wrapped_line_count(cell.value, width))
        existing_height = ws.row_dimensions[row[0].row].height or 0
        calculated_height = max(22, max_lines * 18 + 8)
        ws.row_dimensions[row[0].row].height = min(409, max(existing_height, calculated_height))


def export_plan(payload, output_path):
    plan = payload.get("plan") or {}
    school = clean(plan.get("schoolName") or payload.get("schoolName") or "")
    area = clean(plan.get("areaName") or "")
    grade = clean(plan.get("grade") or "")
    lesson = clean(plan.get("lessonName") or "Yıllık Plan")
    year = clean(plan.get("year") or "2025-2026")
    teacher = clean(plan.get("teacherName") or "")
    mudur = clean(plan.get("mudurName") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Yıllık Plan"
    ws.sheet_view.showGridLines = False

    widths = [20, 8, 42, 54, 32, 34, 28]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    title = f"{year} EĞİTİM-ÖĞRETİM YILI"
    if school:
        title += f" {tr_upper(school)}"
    
    area_upper = tr_upper(area)
    if area_upper and not area_upper.endswith("ALANI"):
        area_upper += " ALANI"
        
    title += f"\n{area_upper} {tr_upper(grade)} {tr_upper(lesson)} DERSİ ÜNİTELENDİRİLMİŞ YILLIK DERS PLANI"
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].font = Font(bold=True, size=13, color="000000")
    ws.row_dimensions[1].height = 48

    headers = ["Hafta / Tarih", "Saat", "Kazanım", "Konu", "Öğretim Teknikleri", "Araç - Gereç", "Açıklama"]
    ws.append(headers)

    rows = build_rows(plan)
    for row in rows:
        if row["kind"] == "break":
            excel_row = ws.max_row + 1
            ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=7)
            cell = ws.cell(excel_row, 1, tr_upper(row["text"]))
            
            fill = PatternFill("solid", fgColor="FEF3C7")
            for col in range(1, 8):
                c = ws.cell(excel_row, col)
                c.fill = fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if col == 1:
                    c.font = Font(bold=True, color="7A3412")
            continue
        note_value = rich_note_text(row["notes"], row.get("examLabels"))
        ws.append([
            row["weekDate"],
            row["hours"],
            row["outcomes"],
            row["topics"],
            row["methods"],
            row["materials"],
            note_value,
        ])

    # Save max row of actual table before adding footer descriptions & signatures
    table_max_row = ws.max_row

    thin = Side(style="thin", color="9CA3AF")
    thick = Side(style="medium", color="374151")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 1. Default description rows before the signatures
    desc_start = ws.max_row + 1
    
    # Description 1
    ws.merge_cells(start_row=desc_start, start_column=1, end_row=desc_start, end_column=7)
    cell1 = ws.cell(desc_start, 1, "Bu plan Mesleki ve Teknik Eğitim Genel Müdürlüğü ile Talim Terbiye Kurulunun yayınladığı Çerçeve Öğretim Programı ve Ders Bilgi Formlarına göre hazırlanmıştır.")
    cell1.font = Font(name="Arial", size=10, italic=True)
    cell1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[desc_start].height = 20
    
    # Description 2
    ws.merge_cells(start_row=desc_start + 1, start_column=1, end_row=desc_start + 1, end_column=7)
    cell2 = ws.cell(desc_start + 1, 1, "Atatürkçülük konuları ile ilgili olarak Talim ve Terbiye Kurulu Başkanlığının 2104 ve 2488 sayılı Tebliğler Dergisinden yararlanılmıştır.")
    cell2.font = Font(name="Arial", size=10, italic=True)
    cell2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[desc_start + 1].height = 20
    
    # 2. Zümre Teachers Signatures
    raw_teachers = re.split(r'[,/]', teacher) if teacher else []
    teachers_list = [t.strip() for t in raw_teachers if t.strip()]
    
    sig_start_row = desc_start + 6
    current_row = sig_start_row
    
    if teachers_list:
        chunks = [teachers_list[i:i + 3] for i in range(0, len(teachers_list), 3)]
        
        for chunk in chunks:
            ws.row_dimensions[current_row].height = 50
            
            if len(chunk) == 3:
                # Column 1: A-C
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                c1 = ws.cell(current_row, 1, f"{chunk[0]}\nZümre Öğretmeni")
                c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c1.font = Font(bold=True)
                
                # Column 2: D-E
                ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
                c2 = ws.cell(current_row, 4, f"{chunk[1]}\nZümre Öğretmeni")
                c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c2.font = Font(bold=True)
                
                # Column 3: F-G
                ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
                c3 = ws.cell(current_row, 6, f"{chunk[2]}\nZümre Öğretmeni")
                c3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c3.font = Font(bold=True)
                
            elif len(chunk) == 2:
                # Column 1: A-D
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                c1 = ws.cell(current_row, 1, f"{chunk[0]}\nZümre Öğretmeni")
                c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c1.font = Font(bold=True)
                
                # Column 2: E-G
                ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
                c2 = ws.cell(current_row, 5, f"{chunk[1]}\nZümre Öğretmeni")
                c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c2.font = Font(bold=True)
                
            elif len(chunk) == 1:
                # Column 1: A-G
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                c1 = ws.cell(current_row, 1, f"{chunk[0]}\nZümre Öğretmeni")
                c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c1.font = Font(bold=True)
                
            current_row += 1
            
    # 3. Okul Müdürü Signature block
    mudur_start = current_row + 5
    ws.row_dimensions[mudur_start].height = 65
    
    ws.merge_cells(start_row=mudur_start, start_column=1, end_row=mudur_start, end_column=7)
    
    start_date_str = ""
    if plan.get("startDate"):
        dt = parse_date(plan.get("startDate"))
        if dt:
            start_date_str = f"{dt.day:02d}/{dt.month:02d}/{dt.year}\n"
            
    m_cell = ws.cell(mudur_start, 1, f"{start_date_str}Uygundur\n{mudur}\nOkul Müdürü")
    m_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    m_cell.font = Font(bold=True)

    # Apply standard table formatting only to actual table rows (rows 1 to table_max_row)
    for row in ws.iter_rows(min_row=1, max_row=table_max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(top=thick, left=thin, right=thin, bottom=thick)
            if cell.column in (1, 2):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif cell.row > 2:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)

    ws.freeze_panes = "A3"
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    if plan.get("addPageNumbers", True):
        ws.oddFooter.center.text = "Sayfa &P / &N"
        ws.oddFooter.center.size = 9

    autosize_rows(ws)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    with open(input_path, "r", encoding="utf-8-sig") as fh:
        payload = json.load(fh)
    export_plan(payload, output_path)
    print(json.dumps({"output": output_path}, ensure_ascii=False))


if __name__ == "__main__":
    main()
